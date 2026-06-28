# -*- coding: utf-8 -*-
from __future__ import annotations

import base64
import cgi
import datetime as dt
import json
import os
import shutil
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from zipfile import is_zipfile

import openpyxl


ROOT = Path(__file__).resolve().parent
IS_VERCEL = bool(os.environ.get("VERCEL"))
RUNTIME_DIR = Path(os.environ.get("CJ_DASHBOARD_RUNTIME_DIR") or ("/tmp/cj-dashboard" if IS_VERCEL else ROOT))
CS_DASHBOARD_ENV = ROOT.parent.parent / "2026-06-27" / "wkr" / "work" / "smartstore-cs-dashboard" / ".env.local"
PUBLIC_DIR = ROOT / "public"
WORK_DIR = RUNTIME_DIR / "work" / "dashboard"
VENDOR_DIR = ROOT / "work" / "vendor"
UPLOAD_DIR = WORK_DIR / "uploads"
OUTPUT_DIR = RUNTIME_DIR / "outputs"
DATA_DIR = RUNTIME_DIR / "data"
SHIPMENTS_DB = DATA_DIR / "shipments.json"
NAVER_API_BASE = "https://api.commerce.naver.com/external"
NAVER_PLACE_ORDER_ENDPOINT = f"{NAVER_API_BASE}/v1/pay-order/seller/product-orders/confirm"
NAVER_DISPATCH_ENDPOINT = f"{NAVER_API_BASE}/v1/pay-order/seller/product-orders/dispatch"
CJ_DELIVERY_COMPANY_CODE = "CJGLS"

if VENDOR_DIR.exists():
    sys.path.insert(0, str(VENDOR_DIR))


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if key and not os.environ.get(key):
            os.environ[key] = value


load_env_file(CS_DASHBOARD_ENV)
load_env_file(ROOT / ".env.local")

SMARTSTORE_HEADERS = [
    "상품주문번호", "주문번호", "배송속성", "풀필먼트사(주문 기준)", "택배사(주문 기준)",
    "배송방법(구매자 요청)", "배송방법", "택배사", "송장번호", "발송일", "판매채널",
    "구매자명", "구매자ID", "수취인명", "주문상태", "주문세부상태", "수량클레임 여부",
    "결제위치", "결제일", "상품번호", "상품명", "상품종류", "반품안심케어", "멤버십N배송",
    "옵션정보", "옵션관리코드", "수량", "옵션가격", "상품가격", "최종 상품별 할인액",
    "최초 상품별 할인액", "판매자 부담 할인액", "최종 상품별 총 주문금액",
    "최초 상품별 총 주문금액", "사은품", "발주확인일", "발송기한", "발송처리일",
    "송장출력일", "배송비 형태", "배송비 묶음번호", "배송비 유형", "배송비 합계",
    "제주/도서 추가배송비", "배송비 할인액", "판매자 상품코드", "판매자 내부코드1",
    "판매자 내부코드2", "수취인연락처1", "수취인연락처2", "통합배송지", "기본배송지",
    "상세배송지", "구매자연락처", "우편번호", "배송메세지", "출고지", "결제수단",
    "네이버페이 주문관리 수수료", "매출연동 수수료", "정산예정금액", "개인통관고유부호",
    "주문일시", "배송희망일", "구독신청회차", "구독진행회차", "구독배송희망일",
    "배송태그 유형", "출입방법 유형", "출입방법 내용", "수령위치 유형", "수령위치 내용",
]

REQUIRED_UPLOAD_HEADERS = [
    "상품주문번호", "주문번호", "수취인명", "수취인연락처1", "통합배송지", "상품명", "수량",
]


@dataclass
class WorkbookRead:
    path: Path
    original_name: str
    password_removed: bool


def ensure_dirs() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not SHIPMENTS_DB.exists():
        SHIPMENTS_DB.write_text("[]", encoding="utf-8")


def json_response(handler: BaseHTTPRequestHandler, status: int, payload: dict) -> None:
    body = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def safe_filename(name: str) -> str:
    keep = [ch if ch.isalnum() or ch in "._-()[] " else "_" for ch in name]
    cleaned = "".join(keep).strip(" .")
    return cleaned or f"upload_{int(time.time())}.xlsx"


def parse_json_body(handler: BaseHTTPRequestHandler) -> dict:
    size = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(size)
    return json.loads(raw.decode("utf-8")) if raw else {}


def decrypt_with_excel(source: Path, password: str, target: Path) -> None:
    password = password.replace("'", "''")
    source_s = str(source).replace("'", "''")
    target_s = str(target).replace("'", "''")
    script = f"""
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
  $wb = $excel.Workbooks.Open('{source_s}', 0, $false, 5, '{password}')
  $wb.Password = ''
  $wb.WritePassword = ''
  $wb.SaveAs('{target_s}', 51, '', '')
  $wb.Close($false)
}} finally {{
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""
    result = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
        timeout=60,
    )
    if result.returncode != 0 or not target.exists():
        raise RuntimeError((result.stderr or result.stdout or "엑셀 비밀번호 해제에 실패했습니다.").strip())


def prepare_workbook(upload_path: Path, password: str, original_name: str) -> WorkbookRead:
    if is_zipfile(upload_path):
        return WorkbookRead(upload_path, original_name, False)
    if not password:
        raise ValueError("암호화된 엑셀입니다. 비밀번호를 입력해주세요.")
    decrypted = upload_path.with_name(upload_path.stem + "_password_removed.xlsx")
    if decrypted.exists():
        decrypted.unlink()
    decrypt_with_excel(upload_path, password, decrypted)
    if not is_zipfile(decrypted):
        raise ValueError("비밀번호 제거 후에도 xlsx 형식으로 읽을 수 없습니다.")
    return WorkbookRead(decrypted, original_name, True)


def read_sheet(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("엑셀 파일에 데이터가 없습니다.")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return wb, ws, headers, rows


def analyze_excel(path: Path) -> dict:
    _, ws, headers, rows = read_sheet(path)
    index = {header: headers.index(header) for header in headers if header}
    data_rows = [row for row in rows[1:] if any(cell not in (None, "") for cell in row)]
    missing = [header for header in REQUIRED_UPLOAD_HEADERS if header not in index]

    def get(row, header):
        col = index.get(header)
        return row[col] if col is not None and col < len(row) else None

    preview = []
    for row in data_rows[:100]:
        preview.append({
            "productOrderNo": str(get(row, "상품주문번호") or ""),
            "orderNo": str(get(row, "주문번호") or ""),
            "receiver": str(get(row, "수취인명") or ""),
            "phone": str(get(row, "수취인연락처1") or ""),
            "address": str(get(row, "통합배송지") or ""),
            "product": str(get(row, "상품명") or ""),
            "option": str(get(row, "옵션정보") or ""),
            "quantity": get(row, "수량") or "",
            "message": str(get(row, "배송메세지") or ""),
        })

    return {
        "sheetName": ws.title,
        "columns": len(headers),
        "rows": len(data_rows),
        "missing": missing,
        "headers": headers,
        "preview": preview,
    }


def make_cj_upload_from_workbook(source: Path) -> Path:
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output = OUTPUT_DIR / f"smartstore_cj_upload_{timestamp}.xlsx"
    wb = openpyxl.load_workbook(source)
    for sheet in list(wb.worksheets)[1:]:
        wb.remove(sheet)
    if wb.security is not None:
        wb.security.workbookPassword = None
        wb.security.lockStructure = False
    wb.save(output)
    return output


def create_naver_secret_sign(client_id: str, client_secret: str, timestamp_ms: str) -> str:
    try:
        import bcrypt  # type: ignore
    except Exception as exc:
        raise RuntimeError("이 로컬 환경에는 bcrypt 모듈이 없어 자동 토큰 발급을 할 수 없습니다. 네이버 커머스API 센터에서 받은 access token을 직접 입력해주세요.") from exc

    client_secret = client_secret.strip().replace("\\$", "$")
    plain = f"{client_id}_{timestamp_ms}".encode("utf-8")
    salt = client_secret.encode("utf-8")
    try:
        hashed = bcrypt.hashpw(plain, salt)
    except ValueError:
        if client_secret.startswith("$2a$"):
            hashed = bcrypt.hashpw(plain, ("$2b$" + client_secret[4:]).encode("utf-8"))
        else:
            raise
    return base64.b64encode(hashed).decode("utf-8")


def http_json(method: str, url: str, token: str | None = None, body: dict | None = None, form: dict | None = None) -> dict:
    proxy_url = (os.environ.get("NAVER_PROXY_URL") or "").strip()
    if proxy_url and url.startswith("https://api.commerce.naver.com/"):
        return http_json_via_proxy(method, url, body=body, form=form)

    headers = {"Accept": "application/json"}
    data = None
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if form is not None:
        data = urllib.parse.urlencode(form).encode("utf-8")
        headers["Content-Type"] = "application/x-www-form-urlencoded"
    elif body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json; charset=utf-8"

    for attempt in range(3):
        request = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                raw = response.read().decode("utf-8")
                return json.loads(raw) if raw else {}
        except urllib.error.HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            if exc.code == 429 and attempt < 2:
                time.sleep(2 + attempt * 3)
                continue
            if "GW.IP_NOT_ALLOWED" in raw:
                raise RuntimeError(
                    "네이버 커머스API 호출이 IP 제한으로 차단됐습니다. "
                    "커머스API 센터의 애플리케이션 설정에서 현재 서버/PC의 외부 IP를 허용 IP로 등록해야 합니다."
                ) from exc
            raise RuntimeError(f"API 오류 {exc.code}: {raw}") from exc
    raise RuntimeError("네이버 API 호출 재시도에 실패했습니다.")


def proxy_endpoint(path: str) -> str:
    proxy_url = (os.environ.get("NAVER_PROXY_URL") or "").strip().rstrip("/")
    return f"{proxy_url}{path}"


def proxy_headers() -> dict:
    secret = (os.environ.get("NAVER_PROXY_SECRET") or "").strip()
    if not secret:
        raise RuntimeError("NAVER_PROXY_URL은 설정되어 있지만 NAVER_PROXY_SECRET이 없습니다.")
    return {
        "Accept": "application/json",
        "Content-Type": "application/json; charset=utf-8",
        "X-Naver-Proxy-Secret": secret,
    }


def http_json_via_proxy(method: str, url: str, body: dict | None = None, form: dict | None = None) -> dict:
    if form is not None:
        # Token requests are handled by the proxy itself.
        proxy_url = proxy_endpoint("/naver/token")
        payload = {}
    else:
        proxy_url = proxy_endpoint("/naver/request")
        payload = {
            "url": url,
            "method": method,
            "headers": {"Accept": "application/json"},
            "body": json.dumps(body, ensure_ascii=False) if body is not None else None,
        }

    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    for attempt in range(3):
        request = urllib.request.Request(proxy_url, data=data, headers=proxy_headers(), method="POST")
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                raw = response.read().decode("utf-8")
                return json.loads(raw) if raw else {}
        except urllib.error.HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            if exc.code == 429 and attempt < 2:
                time.sleep(2 + attempt * 3)
                continue
            raise RuntimeError(f"네이버 고정 IP 프록시 오류 {exc.code}: {raw}") from exc
    raise RuntimeError("네이버 고정 IP 프록시 호출 재시도에 실패했습니다.")


def get_access_token(client_id: str, client_secret: str) -> str:
    timestamp_ms = str(int(time.time() * 1000))
    secret_sign = create_naver_secret_sign(client_id, client_secret, timestamp_ms)
    payload = {
        "client_id": client_id,
        "timestamp": timestamp_ms,
        "client_secret_sign": secret_sign,
        "grant_type": "client_credentials",
        "type": "SELF",
    }
    result = http_json("POST", f"{NAVER_API_BASE}/v1/oauth2/token", form=payload)
    token = result.get("access_token") or result.get("accessToken")
    if not token:
        raise RuntimeError(f"토큰 응답에서 access_token을 찾지 못했습니다: {result}")
    return token


def nested_get(source: dict, *paths):
    for path in paths:
        current = source
        ok = True
        for key in path.split("."):
            if isinstance(current, dict) and key in current:
                current = current[key]
            else:
                ok = False
                break
        if ok and current not in (None, ""):
            return current
    return None


def display_delivery_attribute(value) -> str:
    mapping = {
        "TODAY": "당일발송",
        "NORMAL": "일반배송",
        "HOPE": "희망일배송",
        "ARRIVAL_GUARANTEE": "도착보장",
    }
    return mapping.get(str(value or ""), str(value or ""))


def display_delivery_method(value) -> str:
    mapping = {
        "DELIVERY": "택배,등기,소포",
        "DIRECT_DELIVERY": "직접배송",
        "VISIT_RECEIPT": "방문수령",
        "QUICK_SVC": "퀵서비스",
        "NOTHING": "",
    }
    return mapping.get(str(value or ""), str(value or "택배,등기,소포"))


def collect_product_order_ids(changed_payload: dict, statuses: set[str] | None = None) -> list[str]:
    candidates = (
        changed_payload.get("data")
        or changed_payload.get("lastChangeStatuses")
        or changed_payload.get("contents")
        or changed_payload.get("content")
        or []
    )
    ids = []
    if isinstance(candidates, dict):
        candidates = (
            candidates.get("lastChangeStatuses")
            or candidates.get("contents")
            or candidates.get("content")
            or candidates.get("data")
            or []
        )
    for item in candidates:
        if not isinstance(item, dict):
            continue
        if statuses and item.get("productOrderStatus") not in statuses:
            continue
        value = item.get("productOrderId") or item.get("productOrderNo") or item.get("상품주문번호")
        if value:
            ids.append(str(value))
    return list(dict.fromkeys(ids))


def fetch_naver_orders(payload: dict) -> dict:
    token = (payload.get("accessToken") or "").strip()
    if not token:
        client_id = (payload.get("clientId") or os.environ.get("NAVER_COMMERCE_CLIENT_ID") or "").strip()
        client_secret = (payload.get("clientSecret") or os.environ.get("NAVER_COMMERCE_CLIENT_SECRET") or "").strip()
        if not client_id or not client_secret:
            raise ValueError("access token 또는 client_id/client_secret이 필요합니다.")
        token = get_access_token(client_id, client_secret)

    max_orders = int(payload.get("maxOrders") or 300)
    date_from = payload.get("dateFrom") or dt.datetime.now().strftime("%Y-%m-%d")
    date_to = payload.get("dateTo") or date_from
    start_date = dt.date.fromisoformat(date_from)
    end_date = dt.date.fromisoformat(date_to)
    if end_date < start_date:
        raise ValueError("조회 종료일이 시작일보다 빠릅니다.")

    product_order_ids = []
    changed_keys = set()
    shippable_statuses = {"PAYED"}
    current = start_date
    while current <= end_date and len(product_order_ids) < max_orders:
        day = current.isoformat()
        query = urllib.parse.urlencode({
            "lastChangedFrom": f"{day}T00:00:00.000+09:00",
            "lastChangedTo": f"{day}T23:59:59.999+09:00",
        })
        changed = http_json(
            "GET",
            f"{NAVER_API_BASE}/v1/pay-order/seller/product-orders/last-changed-statuses?{query}",
            token=token,
        )
        changed_keys.update(changed.keys())
        product_order_ids.extend(collect_product_order_ids(changed, shippable_statuses))
        product_order_ids = list(dict.fromkeys(product_order_ids))[:max_orders]
        current += dt.timedelta(days=1)
        time.sleep(0.35)

    product_order_ids = product_order_ids[:max_orders]
    details = []
    for start in range(0, len(product_order_ids), 300):
        batch = product_order_ids[start:start + 300]
        if not batch:
            continue
        detail = http_json(
            "POST",
            f"{NAVER_API_BASE}/v1/pay-order/seller/product-orders/query",
            token=token,
            body={"productOrderIds": batch},
        )
        block = detail.get("data") or detail.get("productOrders") or detail.get("contents") or []
        if isinstance(block, dict):
            block = block.get("contents") or block.get("productOrders") or block.get("data") or []
        if isinstance(block, list):
            details.extend(block)

    rows = [naver_detail_to_row(item) for item in details]
    confirmed_rows = [row for row in rows if is_place_order_confirmed(row)]
    new_rows = [row for row in rows if not is_place_order_confirmed(row)]
    output = create_smartstore_excel(confirmed_rows) if confirmed_rows else None
    return {
        "tokenIssued": not bool(payload.get("accessToken")),
        "changedCount": len(product_order_ids),
        "rows": len(rows),
        "newRows": len(new_rows),
        "confirmedRows": len(confirmed_rows),
        "preview": new_rows[:100],
        "confirmedPreview": confirmed_rows[:100],
        "downloadRows": len(confirmed_rows),
        "downloadName": output.name if output else "",
        "downloadUrl": f"/download/{urllib.parse.quote(output.name)}" if output else "",
        "rawChangedKeys": sorted(changed_keys),
    }


def test_naver_token(payload: dict) -> dict:
    token = resolve_access_token(payload)
    today = dt.datetime.now().strftime("%Y-%m-%d")
    query = urllib.parse.urlencode({
        "lastChangedFrom": f"{today}T00:00:00.000+09:00",
        "lastChangedTo": f"{today}T23:59:59.999+09:00",
    })
    response = http_json(
        "GET",
        f"{NAVER_API_BASE}/v1/pay-order/seller/product-orders/last-changed-statuses?{query}",
        token=token,
    )
    ids = collect_product_order_ids(response)
    return {
        "accessToken": token,
        "proxyEnabled": bool((os.environ.get("NAVER_PROXY_URL") or "").strip()),
        "todayChangedCount": len(ids),
        "rawKeys": list(response.keys()),
    }


def is_place_order_confirmed(row: dict) -> bool:
    value = (
        row.get("발주확인일")
        or row.get("placeOrderDate")
        or row.get("confirmDate")
        or row.get("?諛쒖＜?뺤씤??")
    )
    return bool(str(value or "").strip())


def naver_detail_to_row(item: dict) -> dict:
    product_order = item.get("productOrder") or item.get("productOrderInfo") or item
    order = item.get("order") or item.get("orderInfo") or {}
    delivery = item.get("delivery") or item.get("deliveryInfo") or product_order.get("delivery") or product_order
    shipping = (
        item.get("shippingAddress")
        or item.get("deliveryAddress")
        or product_order.get("shippingAddress")
        or delivery.get("shippingAddress")
        or {}
    )
    product = item.get("product") or item.get("productInfo") or product_order.get("product") or product_order

    receiver = nested_get(shipping, "name", "receiverName", "baseAddress.name") or nested_get(product_order, "receiverName")
    phone1 = nested_get(shipping, "tel1", "phoneNumber1", "receiverTel1", "baseAddress.tel1") or nested_get(product_order, "receiverTel1")
    phone2 = nested_get(shipping, "tel2", "phoneNumber2", "receiverTel2") or ""
    base_addr = nested_get(shipping, "baseAddress", "address1", "receiverAddress1") or ""
    detail_addr = nested_get(shipping, "detailedAddress", "address2", "receiverAddress2") or ""
    full_addr = nested_get(shipping, "fullAddress", "address") or " ".join([str(base_addr), str(detail_addr)]).strip()

    row = {header: "" for header in SMARTSTORE_HEADERS}
    row.update({
        "상품주문번호": str(nested_get(product_order, "productOrderId", "productOrderNo") or ""),
        "주문번호": str(nested_get(order, "orderId", "orderNo") or nested_get(product_order, "orderId", "orderNo") or ""),
        "배송속성": display_delivery_attribute(nested_get(product_order, "deliveryAttributeType", "deliveryAttribute")),
        "택배사(주문 기준)": "CJ대한통운",
        "배송방법(구매자 요청)": display_delivery_method(nested_get(delivery, "deliveryMethod", "deliveryMethodType", "expectedDeliveryMethod")),
        "배송방법": display_delivery_method(nested_get(delivery, "deliveryMethod", "deliveryMethodType", "expectedDeliveryMethod")),
        "판매채널": "스마트스토어",
        "구매자명": nested_get(order, "ordererName", "orderer.name") or "",
        "구매자ID": nested_get(order, "ordererId", "orderer.id") or "",
        "수취인명": receiver or "",
        "주문상태": nested_get(product_order, "productOrderStatus", "orderStatus") or "",
        "주문세부상태": nested_get(product_order, "claimStatus", "claimType") or "",
        "수량클레임 여부": "N",
        "결제위치": nested_get(order, "payLocationType", "paymentLocation") or "",
        "결제일": nested_get(order, "paymentDate", "payDate") or "",
        "상품번호": str(nested_get(product, "productId", "originProductNo") or nested_get(product_order, "productId") or ""),
        "상품명": nested_get(product, "productName", "name") or nested_get(product_order, "productName") or "",
        "상품종류": nested_get(product, "productClass", "productType") or "",
        "옵션정보": nested_get(product_order, "optionInfo", "optionName", "productOption") or "",
        "옵션관리코드": nested_get(product_order, "optionManageCode", "optionCode") or "",
        "수량": nested_get(product_order, "quantity", "orderQuantity") or 1,
        "상품가격": nested_get(product_order, "productPrice", "unitPrice") or "",
        "최종 상품별 할인액": nested_get(product_order, "productDiscountAmount") or "",
        "최초 상품별 할인액": nested_get(product_order, "initialProductDiscountAmount") or "",
        "판매자 부담 할인액": nested_get(product_order, "sellerBurdenStoreDiscountAmount") or "",
        "최종 상품별 총 주문금액": nested_get(product_order, "totalPaymentAmount") or "",
        "최초 상품별 총 주문금액": nested_get(product_order, "initialPaymentAmount") or "",
        "발주확인일": nested_get(product_order, "placeOrderDate", "confirmDate") or "",
        "발송기한": nested_get(product_order, "shippingDueDate", "dispatchDueDate") or "",
        "배송비 형태": nested_get(delivery, "shippingFeeType", "deliveryFeeType") or "",
        "배송비 묶음번호": nested_get(delivery, "packageNumber", "deliveryBundleGroupNo") or "",
        "배송비 유형": nested_get(delivery, "deliveryPolicyType", "deliveryFeePayType") or "",
        "배송비 합계": nested_get(delivery, "deliveryFeeAmount") or "",
        "배송비 할인액": nested_get(delivery, "deliveryDiscountAmount") or "",
        "판매자 상품코드": nested_get(product_order, "sellerProductCode", "merchantProductNo") or "",
        "판매자 내부코드1": nested_get(product_order, "sellerCustomCode1") or "",
        "판매자 내부코드2": nested_get(product_order, "sellerCustomCode2") or "",
        "수취인연락처1": phone1 or "",
        "수취인연락처2": phone2,
        "통합배송지": full_addr,
        "기본배송지": base_addr,
        "상세배송지": detail_addr,
        "구매자연락처": nested_get(order, "ordererTel", "orderer.phone") or "",
        "우편번호": str(nested_get(shipping, "zipCode", "zipcode", "postalCode") or ""),
        "배송메세지": nested_get(delivery, "deliveryMessage", "memo") or nested_get(product_order, "deliveryMessage") or "",
        "출고지": nested_get(product_order, "takingAddress.baseAddress") or "",
        "결제수단": nested_get(order, "paymentMeans", "payMeans") or "",
        "네이버페이 주문관리 수수료": nested_get(product_order, "paymentCommission") or "",
        "매출연동 수수료": nested_get(product_order, "saleCommission") or "",
        "정산예정금액": nested_get(product_order, "expectedSettlementAmount") or "",
        "개인통관고유부호": nested_get(product_order, "individualCustomUniqueCode") or "",
        "주문일시": nested_get(order, "orderDate", "orderYmdt") or "",
        "배송희망일": nested_get(delivery, "hopeDeliveryDate") or "",
    })
    return row


def create_smartstore_excel(rows: list[dict]) -> Path:
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output = OUTPUT_DIR / f"smartstore_api_to_cj_{timestamp}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주발송관리"
    ws.append(SMARTSTORE_HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in SMARTSTORE_HEADERS])

    text_columns = {"상품주문번호", "주문번호", "송장번호", "수취인연락처1", "수취인연락처2", "구매자연락처", "우편번호"}
    for idx, header in enumerate(SMARTSTORE_HEADERS, start=1):
        if header in text_columns:
            for cell in ws.iter_cols(idx, idx, 2, ws.max_row):
                for c in cell:
                    c.number_format = "@"
                    if c.value is not None:
                        c.value = str(c.value)
    ws.freeze_panes = "A2"
    wb.save(output)
    return output


def parse_tracking_file(path: Path) -> dict:
    _, _, headers, rows = read_sheet(path)
    normalized = [str(header or "").replace(" ", "").lower() for header in headers]

    def find_col(candidates: list[str], fallback: int | None = None) -> int | None:
        for idx, header in enumerate(normalized):
            for candidate in candidates:
                if candidate.replace(" ", "").lower() in header:
                    return idx
        if fallback is not None and fallback < len(headers):
            return fallback
        return None

    found = {
        "productOrderNo": find_col(["상품주문번호", "고객주문번호", "주문번호", "productorder"], 0),
        "trackingNo": find_col(["송장번호", "운송장번호", "운송장", "tracking", "invoice"], 1),
        "carrier": find_col(["택배사", "배송사", "운송사", "carrier", "deliverycompany"], 2),
    }
    found = {key: value for key, value in found.items() if value is not None}
    records = []
    if "trackingNo" in found:
        for row in rows[1:]:
            tracking = row[found["trackingNo"]] if found["trackingNo"] < len(row) else None
            if not tracking:
                continue
            order_col = found.get("productOrderNo")
            carrier_col = found.get("carrier")
            records.append({
                "productOrderNo": str(row[order_col] if order_col is not None and order_col < len(row) else ""),
                "trackingNo": str(tracking),
                "carrier": str(row[carrier_col] if carrier_col is not None and carrier_col < len(row) else "CJ대한통운"),
                "deliveryCompanyCode": CJ_DELIVERY_COMPANY_CODE,
                "dispatchDate": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
                "status": "READY",
            })
    return {"headers": headers, "mapping": found, "records": records[:500], "rows": len(records)}


def load_shipments() -> list[dict]:
    ensure_dirs()
    try:
        data = json.loads(SHIPMENTS_DB.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_shipments(records: list[dict]) -> None:
    ensure_dirs()
    SHIPMENTS_DB.write_text(json.dumps(records, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def upsert_shipments(records: list[dict]) -> None:
    existing = load_shipments()
    by_key = {
        f"{item.get('productOrderNo', '')}|{item.get('trackingNo', '')}": item
        for item in existing
    }
    now = dt.datetime.now().isoformat(timespec="seconds")
    for record in records:
        key = f"{record.get('productOrderNo', '')}|{record.get('trackingNo', '')}"
        merged = {**by_key.get(key, {}), **record}
        merged.setdefault("createdAt", now)
        merged["updatedAt"] = now
        by_key[key] = merged
    save_shipments(list(by_key.values()))


def filter_shipments(params: dict) -> dict:
    records = load_shipments()
    date_from = params.get("dateFrom") or ""
    date_to = params.get("dateTo") or ""
    query = (params.get("query") or "").strip()
    status = (params.get("status") or "").strip()

    def include(item: dict) -> bool:
        date_value = str(item.get("dispatchDate") or item.get("createdAt") or "")[:10]
        if date_from and date_value < date_from:
            return False
        if date_to and date_value > date_to:
            return False
        if status and item.get("status") != status:
            return False
        if query:
            text = " ".join(str(item.get(key, "")) for key in ("productOrderNo", "trackingNo", "carrier", "receiver", "product"))
            if query not in text:
                return False
        return True

    filtered = [item for item in records if include(item)]
    summary = {
        "total": len(filtered),
        "ready": sum(1 for item in filtered if item.get("status") == "READY"),
        "sent": sum(1 for item in filtered if item.get("status") == "SENT"),
        "failed": sum(1 for item in filtered if item.get("status") == "FAILED"),
    }
    return {"rows": filtered[-1000:], "summary": summary}


def resolve_access_token(payload: dict) -> str:
    token = (payload.get("accessToken") or "").strip()
    if token:
        return token
    client_id = (payload.get("clientId") or os.environ.get("NAVER_COMMERCE_CLIENT_ID") or "").strip()
    client_secret = (payload.get("clientSecret") or os.environ.get("NAVER_COMMERCE_CLIENT_SECRET") or "").strip()
    if not client_id or not client_secret:
        raise ValueError("access token 또는 client_id/client_secret이 필요합니다.")
    return get_access_token(client_id, client_secret)


def dispatch_payload(records: list[dict]) -> dict:
    dispatches = []
    for record in records:
        product_order_no = str(record.get("productOrderNo") or "").strip()
        tracking_no = str(record.get("trackingNo") or "").strip()
        if not product_order_no or not tracking_no:
            continue
        dispatches.append({
            "productOrderId": product_order_no,
            "deliveryMethod": "DELIVERY",
            "deliveryCompanyCode": record.get("deliveryCompanyCode") or CJ_DELIVERY_COMPANY_CODE,
            "trackingNumber": tracking_no,
            "dispatchDate": record.get("dispatchDate") or dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        })
    return {"dispatchProductOrders": dispatches}


def product_order_ids_from_payload(payload: dict) -> list[str]:
    records = payload.get("records") or []
    ids = payload.get("productOrderIds") or []
    collected = []
    if isinstance(ids, list):
        collected.extend(str(value).strip() for value in ids if str(value).strip())
    if isinstance(records, list):
        for record in records:
            if not isinstance(record, dict):
                continue
            value = (
                record.get("productOrderNo")
                or record.get("productOrderId")
                or record.get("상품주문번호")
                or record.get("?곹뭹二쇰Ц踰덊샇")
            )
            if value:
                collected.append(str(value).strip())
    return list(dict.fromkeys(value for value in collected if value))


def place_orders(payload: dict) -> dict:
    token = resolve_access_token(payload)
    records = payload.get("records") or []
    product_order_ids = product_order_ids_from_payload(payload)
    if not product_order_ids:
        raise ValueError("발주확인할 주문 목록이 없습니다.")

    results = []
    confirmed = 0
    failed = 0
    confirmed_records = []
    record_by_id = {
        str(record.get("productOrderNo") or record.get("productOrderId") or record.get("상품주문번호") or "").strip(): record
        for record in records
        if isinstance(record, dict)
    }
    for start in range(0, len(product_order_ids), 30):
        batch = product_order_ids[start:start + 30]
        try:
            response = http_json(
                "POST",
                NAVER_PLACE_ORDER_ENDPOINT,
                token=token,
                body={"productOrderIds": batch},
            )
            confirmed += len(batch)
            confirmed_records.extend(record_by_id.get(order_id, {"productOrderNo": order_id}) for order_id in batch)
            results.append({"ok": True, "count": len(batch), "response": response})
        except Exception as exc:
            failed += len(batch)
            results.append({"ok": False, "count": len(batch), "error": str(exc)})
    output = create_smartstore_excel(confirmed_records) if confirmed_records and failed == 0 else None
    return {
        "confirmed": confirmed,
        "failed": failed,
        "results": results,
        "downloadRows": len(confirmed_records) if output else 0,
        "downloadName": output.name if output else "",
        "downloadUrl": f"/download/{urllib.parse.quote(output.name)}" if output else "",
    }


def dispatch_to_smartstore(payload: dict) -> dict:
    token = resolve_access_token(payload)
    records = payload.get("records") or []
    if not isinstance(records, list) or not records:
        raise ValueError("전송할 송장 목록이 없습니다.")

    results = []
    sent_records = []
    for start in range(0, len(records), 30):
        batch = records[start:start + 30]
        body = dispatch_payload(batch)
        if not body["dispatchProductOrders"]:
            continue
        try:
            response = http_json("POST", NAVER_DISPATCH_ENDPOINT, token=token, body=body)
            for record in batch:
                updated = {**record, "status": "SENT", "sentAt": dt.datetime.now().isoformat(timespec="seconds"), "response": response}
                sent_records.append(updated)
            results.append({"ok": True, "count": len(batch), "response": response})
        except Exception as exc:
            for record in batch:
                sent_records.append({**record, "status": "FAILED", "error": str(exc), "updatedAt": dt.datetime.now().isoformat(timespec="seconds")})
            results.append({"ok": False, "count": len(batch), "error": str(exc)})

    upsert_shipments(sent_records)
    return {
        "sent": sum(item.get("status") == "SENT" for item in sent_records),
        "failed": sum(item.get("status") == "FAILED" for item in sent_records),
        "results": results,
        "records": sent_records,
    }


class DashboardHandler(BaseHTTPRequestHandler):
    server_version = "CJSmartstoreDashboard/0.2"

    def log_message(self, format: str, *args) -> None:
        return

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path in ("/", "/index.html"):
            self.serve_file(PUBLIC_DIR / "index.html", "text/html; charset=utf-8")
        elif parsed.path == "/styles.css":
            self.serve_file(PUBLIC_DIR / "styles.css", "text/css; charset=utf-8")
        elif parsed.path == "/app.js":
            self.serve_file(PUBLIC_DIR / "app.js", "application/javascript; charset=utf-8")
        elif parsed.path.startswith("/download/"):
            name = urllib.parse.unquote(parsed.path.replace("/download/", "", 1))
            target = OUTPUT_DIR / name
            if not target.exists() or target.parent.resolve() != OUTPUT_DIR.resolve():
                self.send_error(404)
                return
            self.serve_file(target, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            self.send_error(404)

    def serve_file(self, path: Path, content_type: str) -> None:
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self) -> None:
        ensure_dirs()
        path = urllib.parse.urlparse(self.path).path
        routes = {
            "/api/orders/fetch": self.handle_fetch_orders,
            "/api/orders/place-order": self.handle_place_orders,
            "/api/token/test": self.handle_token_test,
            "/api/smartstore/upload": self.handle_excel_upload,
            "/api/tracking/upload": self.handle_tracking_upload,
            "/api/dispatch/send": self.handle_dispatch_send,
            "/api/shipments/list": self.handle_shipments_list,
        }
        handler = routes.get(path)
        if handler:
            handler()
            return
        if path.startswith("/api/"):
            json_response(self, 404, {"ok": False, "error": f"지원하지 않는 API 경로입니다: {path}"})
            return
        self.send_error(404)

    def parse_upload(self):
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
            },
        )
        file_item = form["file"] if "file" in form else None
        if file_item is None or not file_item.filename:
            raise ValueError("파일을 선택해주세요.")
        filename = safe_filename(Path(file_item.filename).name)
        saved = UPLOAD_DIR / f"{int(time.time())}_{filename}"
        with saved.open("wb") as out:
            shutil.copyfileobj(file_item.file, out)
        return saved, form.getfirst("password", "") or "", filename

    def handle_fetch_orders(self) -> None:
        try:
            payload = parse_json_body(self)
            result = fetch_naver_orders(payload)
            json_response(self, 200, {"ok": True, **result})
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def handle_place_orders(self) -> None:
        try:
            payload = parse_json_body(self)
            result = place_orders(payload)
            json_response(self, 200, {"ok": True, **result})
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def handle_token_test(self) -> None:
        try:
            payload = parse_json_body(self)
            result = test_naver_token(payload)
            json_response(self, 200, {"ok": True, **result})
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def handle_excel_upload(self) -> None:
        try:
            upload, password, original = self.parse_upload()
            prepared = prepare_workbook(upload, password, original)
            analysis = analyze_excel(prepared.path)
            if analysis["missing"]:
                json_response(self, 422, {"ok": False, "error": "필수 컬럼이 없습니다.", **analysis})
                return
            output = make_cj_upload_from_workbook(prepared.path)
            json_response(self, 200, {
                "ok": True,
                "passwordRemoved": prepared.password_removed,
                "downloadName": output.name,
                "downloadUrl": f"/download/{urllib.parse.quote(output.name)}",
                **analysis,
            })
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def handle_tracking_upload(self) -> None:
        try:
            upload, password, original = self.parse_upload()
            prepared = prepare_workbook(upload, password, original)
            parsed = parse_tracking_file(prepared.path)
            upsert_shipments(parsed.get("records", []))
            json_response(self, 200, {"ok": True, "passwordRemoved": prepared.password_removed, **parsed})
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def handle_dispatch_send(self) -> None:
        try:
            payload = parse_json_body(self)
            result = dispatch_to_smartstore(payload)
            json_response(self, 200, {"ok": True, **result})
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})

    def handle_shipments_list(self) -> None:
        try:
            payload = parse_json_body(self)
            result = filter_shipments(payload)
            json_response(self, 200, {"ok": True, **result})
        except Exception as exc:
            json_response(self, 400, {"ok": False, "error": str(exc)})


def main() -> None:
    ensure_dirs()
    port = int(os.environ.get("PORT", "8765"))
    server = ThreadingHTTPServer(("127.0.0.1", port), DashboardHandler)
    print(f"dashboard=http://127.0.0.1:{port}", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass


if __name__ == "__main__":
    main()
